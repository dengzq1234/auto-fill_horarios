#!/usr/bin/python
# Ziqi DENG 2021

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image

from calendar import monthrange
from holidays_es import Province
import itertools

import pandas as pd
import datetime
from datetime import date, timedelta
import os
import sys

"""
DEPENDENCIES:
pandas
openpyxl
holidays-es
pillow
BeautifulSoup4
"""
# fill basic info
NAME='ZIQI DENG'
NIF='EDXXXXXX'
YEAR=2020
HORA_ENTRADA_MORNING = "09:00:00"
HORA_SALIDA_MORNING = "13:00:00"
HORA_ENTRADA_AFTERNOON = "14:00:00"
HORA_SALIDA_AFTERNOON = "17:30:00"
TOTAL_HORAS = 7.5

ANNUAL_LEAVES = [
    # start-date, end date Year-Month-Day
    ["2020-01-04", "2020-01-10"], 
    ["2020-02-04", "2020-02-14"],
]
TEMPLATE = './registro_jornada_laboral_template.xlsx'

#month-dictionary in spanish
months_list = [
        "enero",
        "febrero",
        "marzo",
        "abril",
        "mayo",
        "junio",
        "julio",
        "agosto",
        "septiembre",
        "octubre",
        "noviembre",
        "diciembre",
    ]



def main(month):
    wb = load_workbook(TEMPLATE)
    MES=months_list[month-1].upper()
    
    ws = wb.get_sheet_by_name('Hoja1') #Getting the sheet named as 'data'
    ws.cell(row=9, column=2).value = NAME
    ws.cell(row=10, column=2).value = NIF
    ws.cell(row=12, column=2).value = YEAR
    ws.cell(row=13, column=2).value = MES

    # load spanish holidays
    holidays = list(itertools.chain.from_iterable(list(Province(name="madrid", year=YEAR).holidays().values())))

    # load personal holidays
    personal_holidays = []

    for leave in ANNUAL_LEAVES:
        
        sdate = datetime.datetime.strptime(leave[0], "%Y-%m-%d").date()   # start date
        edate = datetime.datetime.strptime(leave[1], "%Y-%m-%d").date()   # start date
        delta = edate - sdate       # as timedelta

        for i in range(delta.days + 1):
            day = sdate + timedelta(days=i)
            personal_holidays.append(day)

    # styling
    from openpyxl.styles import Border, Side

    border_type=Side(border_style=None, color='FF000000')
    border = Border(left=border_type,
                    right=border_type,
                    top=border_type,
                    bottom=border_type,
                    diagonal=border_type,
                    diagonal_direction=0,
                    outline=border_type,
                    vertical=border_type,
                    horizontal=border_type
    )
    # Style 1 - narrow border, black
    thin = Side(border_style="thin", color="000000")#Border style, color
    border = Border(left=thin, right=thin, top=thin, bottom=thin)#Position of border

    ws['A6'].border = border #A6 cell setting border

    for row in ws['A8:B13']:
        for cell in row:
            cell.border = border#A8:G13 range cell setting border

    for row in ws['A15:I47']:
        for cell in row:
            cell.border = border#A8:G13 range cell setting border

    # fill the date info
    days_of_month = monthrange(YEAR, month)[1]
    for day in range(1, days_of_month+1):
        date_of_today = datetime.date(YEAR, month, day)
        # fill the gap
        START_ROW = 15
        ws.cell(row=START_ROW+day, column=1).value = day

    #     #check if is holiday
    #     if date_of_today in holidays:
    #         # fill INCIDENCIA
    #         ws.cell(row=START_ROW+day, column=9).value = "Festivo" 
    #         pass
            
        
        if date_of_today in personal_holidays:
            ws.cell(row=START_ROW+day, column=9).value = "Vacaciones"
            continue
        else:
            pass
        
        if date_of_today in holidays:
            # fill INCIDENCIA
            ws.cell(row=START_ROW+day, column=9).value = "Festivo" 
            continue
        else:
            pass
    
        
        # check if is weekend
        if date_of_today.weekday() == 0 or date_of_today.weekday() == 6:
            continue
        else:
            ws.cell(row=START_ROW+day, column=2).value = HORA_ENTRADA_MORNING
            ws.cell(row=START_ROW+day, column=3).value = HORA_SALIDA_MORNING
            ws.cell(row=START_ROW+day, column=4).value = HORA_ENTRADA_AFTERNOON
            ws.cell(row=START_ROW+day, column=5).value = HORA_SALIDA_AFTERNOON
            ws.cell(row=START_ROW+day, column=8).value = TOTAL_HORAS

    output_file = './registro_jornada_laboral_{}.xlsx'.format(MES+str(YEAR))
    wb.save(output_file)
    return

for month in range(1, 13): # for 12 months
    main(month)

    